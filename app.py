from flask import Flask, request, jsonify
from flask_cors import CORS
import os
from dotenv import load_dotenv
import traceback
import openai
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook

load_dotenv()

app = Flask(__name__)

# ✅ Allow CORS from Vercel frontend
CORS(app, resources={r"/generate-response": {"origins": ["https://www.acuminds.com"]}})

api_key = os.getenv("OPENAI_API_KEY")
client = openai.OpenAI(api_key=api_key)

uploaded_dataset_context = ""

@app.route('/generate-response', methods=['POST'])
def generate_response():
    global uploaded_dataset_context
    try:
        query = request.form.get('query', '')
        file = request.files.get('file')
        if not query:
            return jsonify({"error": "Query is required."}), 400

        context = ""
        if file:
            try:
                file_content = file.read()
                file_stream = BytesIO(file_content)

                if file.filename.endswith(".csv"):
                    df = pd.read_csv(file_stream)
                else:
                    try:
                        workbook = load_workbook(filename=file_stream, data_only=True)
                        sheet = workbook.active
                        data = [row for row in sheet.iter_rows(values_only=True)]
                        df = pd.DataFrame(data[1:], columns=data[0])
                    except TypeError:
                        return jsonify({"error": "Excel file is malformed. Please re-save or upload as .csv."}), 400

                row_count, col_count = df.shape
                col_names = df.columns.tolist()
                data_types = df.dtypes.astype(str).to_dict()
                head_sample = df.head(5).to_dict(orient="records")
                describe_summary = df.describe(include='all').fillna("N/A").to_dict()

                context = (
                    f"You are helping the user analyze a dataset with {row_count} rows and {col_count} columns.\n"
                    f"Column names: {', '.join(col_names)}\n"
                    f"Column types: {data_types}\n"
                    f"Stats: {describe_summary}\n"
                    f"First 5 rows: {head_sample}\n"
                    f"Use this only if the question is related to the dataset."
                )
                uploaded_dataset_context = context

            except Exception as e:
                return jsonify({"error": f"File processing failed: {str(e)}"}), 400

        # Compose message for GPT
        messages = []
        if uploaded_dataset_context:
            messages.append({"role": "system", "content": uploaded_dataset_context})
        messages.append({"role": "user", "content": query.strip()})

        response = client.chat.completions.create(
            model="gpt-3.5-turbo",
            messages=messages
        )

        answer = response.choices[0].message.content.strip()
        return jsonify({"response": answer})

    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.route('/reset-context', methods=['POST'])
def reset_context():
    global uploaded_dataset_context
    uploaded_dataset_context = ""
    return jsonify({"message": "Context reset."})

if __name__ == '__main__':
    app.run(debug=True)